import logging
import datetime
from typing import Optional, Dict, Any
from django.core.management.base import BaseCommand
from django.db import transaction
from core.models import Cadete, Atendimento, Profissional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def parse_bool(v: Any) -> bool:
    """Converte valores variados para booleano."""
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip().upper() in {"X", "SIM", "S", "TRUE", "1"}
    if isinstance(v, (int, float)):
        return bool(v)
    return bool(v)


def parse_date(v: Any) -> Optional[datetime.date]:
    """Parse de data com tratamento robusto de erros."""
    if v is None:
        return None
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return None
        # Tenta múltiplos formatos
        for fmt in ["%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%Y-%m-%d"]:
            try:
                return datetime.datetime.strptime(v, fmt).date()
            except ValueError:
                continue
        logger.warning(f"Não foi possível converter data: '{v}'")
        return None
    return None


def parse_time(v: Any) -> Optional[datetime.time]:
    """Parse de hora."""
    if isinstance(v, datetime.time):
        return v
    if isinstance(v, datetime.datetime):
        return v.time()
    return None


class Command(BaseCommand):
    help = "Importa dados da planilha projeto_saude.xlsm (abas BD_Cadetes, BD_Atendimento, BD_Profissionais)"

    def add_arguments(self, parser):
        parser.add_argument("--arquivo", default="projeto_saude.xlsm")
        parser.add_argument("--skip-profissionais", action="store_true", help="Pular importação de profissionais")
        parser.add_argument("--skip-cadetes", action="store_true", help="Pular importação de cadetes")
        parser.add_argument("--skip-atendimentos", action="store_true", help="Pular importação de atendimentos")

    @transaction.atomic
    def handle(self, *args, **opts):
        """Handle com suporte a rollback em caso de erro."""
        arquivo = opts["arquivo"]
        
        try:
            wb = load_workbook(arquivo, data_only=True, keep_vba=True)
        except Exception as e:
            self.stdout.write(self.style.ERROR(f"Erro ao abrir arquivo: {e}"))
            return

        # Profissionais
        if not opts["skip_profissionais"] and "BD_Profissionais" in wb.sheetnames:
            self._importar_profissionais(wb["BD_Profissionais"])
            
        # Cadetes
        if not opts["skip_cadetes"] and "BD_Cadetes" in wb.sheetnames:
            self._importar_cadetes(wb["BD_Cadetes"])
            
        # Atendimentos
        if not opts["skip_atendimentos"] and "BD_Atendimento" in wb.sheetnames:
            self._importar_atendimentos(wb["BD_Atendimento"])

        self.stdout.write(self.style.SUCCESS("Importação concluída com sucesso!"))

    def _importar_profissionais(self, ws) -> None:
        """Importa profissionais com tratamento de erros por linha."""
        tipos = [ws.cell(1, c).value for c in range(1, ws.max_column + 1) if ws.cell(1, c).value]
        map_titulo = {
            "Médico": "medico",
            "Fisioterapeuta": "fisioterapeuta",
            "Educador Físico": "educador_fisico",
            "Nutricionista": "nutricionista",
            "Psicopedagogo": "psicopedagogo",
            "S-RED": "sred",
        }
        
        criados = 0
        for col, titulo in enumerate(tipos, start=1):
            tipo = map_titulo.get(titulo)
            if not tipo:
                continue
            for r in range(2, ws.max_row + 1):
                try:
                    ident = ws.cell(r, col).value
                    if ident:
                        obj, created = Profissional.objects.get_or_create(
                            tipo=tipo,
                            identificador=str(ident).strip()
                        )
                        if created:
                            criados += 1
                except Exception as e:
                    logger.error(f"Erro ao importar profissional linha {r}: {e}")
                    
        self.stdout.write(self.style.SUCCESS(f"✓ Profissionais: {criados} criados"))

    def _importar_cadetes(self, ws) -> None:
        """Importa cadetes com tratamento de erros por linha."""
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}
        
        criados = 0
        atualizados = 0
        
        for r in range(2, ws.max_row + 1):
            try:
                numero = ws.cell(r, idx.get("Número")).value
                if not numero:
                    continue
                    
                defaults = {
                    "nome": ws.cell(r, idx.get("Nome")).value or f"Cadete {numero}",
                    "nome_de_guerra": ws.cell(r, idx.get("Nome de Guerra")).value,
                    "curso": ws.cell(r, idx.get("Curso")).value,
                    "ano": ws.cell(r, idx.get("Ano")).value,
                    "subunidade": ws.cell(r, idx.get("Subunidade")).value,
                    "pelotao": ws.cell(r, idx.get("Pelotão")).value,
                    "cmt_curso": ws.cell(r, idx.get("Cmt Curso")).value,
                    "cmt_subunidade": ws.cell(r, idx.get("Cmt Subunidade")).value,
                    "cmt_pelotao": ws.cell(r, idx.get("Cmt Pelotão")).value,
                }
                obj, created = Cadete.objects.update_or_create(numero=int(numero), defaults=defaults)
                if created:
                    criados += 1
                else:
                    atualizados += 1
            except Exception as e:
                logger.error(f"Erro ao importar cadete linha {r}: {e}")
                
        self.stdout.write(self.style.SUCCESS(f"✓ Cadetes: {criados} criados, {atualizados} atualizados"))

    def _importar_atendimentos(self, ws) -> None:
        """Importa atendimentos com tratamento de erros por linha."""
        headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
        idx = {h: i + 1 for i, h in enumerate(headers) if h}

        criados = 0
        erros = 0
        
        for r in range(3, ws.max_row + 1):
            try:
                nr = ws.cell(r, idx.get("Nr Cadete")).value
                if not nr:
                    continue
                    
                # Usar get_or_create para evitar race condition
                cadete, created = Cadete.objects.get_or_create(
                    numero=int(nr),
                    defaults={"nome": f"Cadete {nr}"}
                )

                data = parse_date(ws.cell(r, idx.get("Data")).value)
                if not data:
                    logger.warning(f"Atendimento linha {r}: data inválida")
                    erros += 1
                    continue

                atendimento_txt = (ws.cell(r, idx.get("Atendimento")).value or "").strip().lower()
                atendimento = "retorno" if "ret" in atendimento_txt else "inicial"

                Atendimento.objects.create(
                    cadete=cadete,
                    data=data,
                    hora=parse_time(ws.cell(r, idx.get("Hora")).value),
                    atendimento=atendimento,
                    profissional_tipo=ws.cell(r, idx.get("Profissional")).value,
                    profissional_nome=ws.cell(r, idx.get("Nome Profissional")).value,
                    lesao_tipo=ws.cell(r, idx.get("Lesão")).value,
                    parte_do_corpo=ws.cell(r, idx.get("Parte do Corpo")).value,
                    lateralidade=ws.cell(r, idx.get("Lateralidade")).value,
                    parte_lesionada=ws.cell(r, idx.get("Parte Lesionada")).value,
                    local_da_lesao=ws.cell(r, idx.get("Local da Lesão")).value,
                    origem_da_lesao=ws.cell(r, idx.get("Origem da Lesão")).value,
                    sred=ws.cell(r, idx.get("S-RED")).value,
                    causa=ws.cell(r, idx.get("Causa")).value,
                    atividade=ws.cell(r, idx.get("Atividade")).value,
                    tfm_taf=ws.cell(r, idx.get("TFM/TAF")).value,
                    modalidade=ws.cell(r, idx.get("Modalidade")).value,
                    tratamento=ws.cell(r, idx.get("Tratamento")).value,
                    medicamentoso=ws.cell(r, idx.get("Medicamentoso")).value,
                    fisioterapia=parse_bool(ws.cell(r, idx.get("Fisioterapia")).value),
                    sef=parse_bool(ws.cell(r, idx.get("SEF")).value),
                    nutricionista=parse_bool(ws.cell(r, idx.get("Nutricionista")).value),
                    psicopedagogica=parse_bool(ws.cell(r, idx.get("Psicopedagógica")).value),
                    rx=parse_bool(ws.cell(r, idx.get("RX")).value),
                    usg=parse_bool(ws.cell(r, idx.get("USG")).value),
                    tc=parse_bool(ws.cell(r, idx.get("TC")).value),
                    rm=parse_bool(ws.cell(r, idx.get("RM")).value),
                    dexa=parse_bool(ws.cell(r, idx.get("DEXA")).value),
                    sangue=parse_bool(ws.cell(r, idx.get("Sangue")).value),
                    dispensa=parse_bool(ws.cell(r, idx.get("Dispensa")).value),
                    vcl=parse_bool(ws.cell(r, idx.get("VCL")).value),
                    alta=parse_bool(ws.cell(r, idx.get("Alta")).value),
                    risco_cirurgico=parse_bool(ws.cell(r, idx.get("Risco Cirúrgico")).value),
                )
                criados += 1
            except Exception as e:
                logger.error(f"Erro ao importar atendimento linha {r}: {e}")
                erros += 1

        self.stdout.write(self.style.SUCCESS(f"✓ Atendimentos: {criados} criados, {erros} erros"))